print('Starting Now')
import time
import os.path
from openpyxl import Workbook, load_workbook

def collatz_steps(n):
    steps = 0
    while n != 1:
        if n % 2 == 0:
            n //= 2
        else:
            n = 3 * n + 1
        steps += 1
    return steps

def write_to_excel(filename, data):
    if not os.path.isfile(filename):
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Number", "Steps"])
    else:
        wb = load_workbook(filename)
        sheet = wb.active
    sheet.append(data)
    print(data[0])
    wb.save(filename)

def collatz_to_excel(filename, start, end):
    directory = os.path.dirname(filename)
    os.makedirs(directory, exist_ok=True)

    for num in range(start, end + 1):
        steps = collatz_steps(num)
        write_to_excel(filename, [num, steps])

        elapsed_time = time.time() - start_time
        elapsed_hours = elapsed_time // 3600
        elapsed_minutes = (elapsed_time % 3600) // 60

        if elapsed_hours >= 5 and elapsed_minutes >= 58:
            print("Time limit reached. Stopping execution.")
            break

    print("All numbers written!")

start = 27_70_000
end = start + 10000 
start_time = time.time()
collatz_to_excel(f"Excels/collatz_steps {start} to {end}.xlsx", start, end)